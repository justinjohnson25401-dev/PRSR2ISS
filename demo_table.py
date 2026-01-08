import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from collections import Counter

INPUT_FILE = "Таблица.xlsx"          # исходный файл
SHEET_NAME = "Компании"               # имя листа
MAX_ROWS_PER_FILE = 1000              # компаний в одном выходном файле

def style_and_links(ws, headers, zone_counts):
    max_row = ws.max_row
    max_col = len(headers)

    # Вставляем статистику по зонам в строку 1 (над таблицей)
    ws.insert_rows(1)
    stat_text = " | ".join([f"{zone}: {count}" 
                            for zone, count in sorted(zone_counts.items())])
    stat_cell = ws.cell(row=1, column=1)
    stat_cell.value = f"📊 Статистика по зонам: {stat_text}"
    stat_cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws.merge_cells(f"A1:{get_column_letter(max_col)}1")
    stat_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # Теперь данные начинаются со строки 2, заголовки в строке 2
    header_row = 2

    # Формат как таблица (начиная с A2)
    tab = Table(displayName="SalonTable",
                ref=f"A{header_row}:{get_column_letter(max_col)}{max_row}")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Заголовки (строка 2)
    header_fill = PatternFill(start_color="D3D3D3",
                              end_color="D3D3D3",
                              fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="000000")
    header_alignment = Alignment(horizontal="center",
                                 vertical="center",
                                 wrap_text=True)

    for col_num in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Закрепляем статистику и заголовки
    ws.freeze_panes = "A3"

    # Общий стиль
    cell_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for row in ws.iter_rows(min_row=header_row + 1, max_row=max_row,
                            min_col=1, max_col=max_col):
        for cell in row:
            cell.font = cell_font
            cell.border = thin_border

    # Определяем индексы нужных столбцов
    col_idx = {ws.cell(row=header_row, column=i).value: i
               for i in range(1, max_col + 1)}

    col_distance = col_idx.get("Расст. от центра (км)")
    col_rating = col_idx.get("Рейтинг")
    col_reviews = col_idx.get("Оценок")
    col_feedback = col_idx.get("Отзывов")
    col_lat = col_idx.get("Широта")
    col_lon = col_idx.get("Долгота")
    col_2gis = col_idx.get("Открыть в 2ГИС")
    col_yandex = col_idx.get("Открыть в Яндекс")
    col_zone = col_idx.get("Зона")

    # Выравнивание и форматы чисел
    for row_num in range(header_row + 1, max_row + 1):
        # Числовые
        for col_num in [col_distance, col_rating,
                        col_reviews, col_feedback,
                        col_lat, col_lon]:
            if col_num:
                ws.cell(row=row_num, column=col_num).alignment = Alignment(
                    horizontal="right", vertical="center"
                )

        # Текстовые
        for header in ["Название", "Категория", "Специализация", "Адрес",
                       "Зона", "Телефоны", "Email", "Сайт", "Telegram",
                       "Telegram username", "VK", "WhatsApp",
                       "Прочие соцсети", "График работы",
                       "Открыть в 2ГИС", "Открыть в Яндекс"]:
            c = col_idx.get(header)
            if c:
                ws.cell(row=row_num, column=c).alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

        # Форматы (только отображение, значения не меняются)
        if col_distance:
            ws.cell(row=row_num, column=col_distance).number_format = "0.0"
        if col_rating:
            ws.cell(row=row_num, column=col_rating).number_format = "0.0"
        if col_reviews:
            ws.cell(row=row_num, column=col_reviews).number_format = "0"
        if col_feedback:
            ws.cell(row=row_num, column=col_feedback).number_format = "0"
        if col_lat:
            ws.cell(row=row_num, column=col_lat).number_format = "0.000000"
        if col_lon:
            ws.cell(row=row_num, column=col_lon).number_format = "0.000000"

        # ГИПЕРССЫЛКИ 2ГИС и Яндекс (по координатам)
        if col_lat and col_lon and col_2gis and col_yandex:
            lat = ws.cell(row=row_num, column=col_lat).value
            lon = ws.cell(row=row_num, column=col_lon).value
            if lat is not None and lon is not None:
                try:
                    lat_f = float(lat)
                    lon_f = float(lon)
                    
                    # 2ГИС
                    url_2gis = f"https://2gis.ru/geo/{lon_f}%2C{lat_f}"
                    cell_2gis = ws.cell(row=row_num, column=col_2gis)
                    cell_2gis.value = "2ГИС"
                    cell_2gis.hyperlink = url_2gis
                    cell_2gis.style = "Hyperlink"

                    # Яндекс
                    url_yandex = f"https://yandex.ru/maps/?pt={lon_f},{lat_f}&z=17&l=map"
                    cell_yandex = ws.cell(row=row_num, column=col_yandex)
                    cell_yandex.value = "Яндекс"
                    cell_yandex.hyperlink = url_yandex
                    cell_yandex.style = "Hyperlink"
                except (ValueError, TypeError):
                    pass

    # Условное форматирование: рейтинг
    if col_rating:
        col_letter = get_column_letter(col_rating)
        color_scale_rating = ColorScaleRule(
            start_type="num", start_value=3.0, start_color="FFC7CE",
            mid_type="num", mid_value=4.2, mid_color="FFEB9C",
            end_type="num", end_value=5.0, end_color="C6EFCE"
        )
        ws.conditional_formatting.add(
            f"{col_letter}{header_row + 1}:{col_letter}{max_row}",
            color_scale_rating
        )

    # Условное форматирование: расстояние
    if col_distance:
        col_letter = get_column_letter(col_distance)
        color_scale_distance = ColorScaleRule(
            start_type="num", start_value=0.5, start_color="C6EFCE",
            mid_type="num", mid_value=7.5, mid_color="FFEB9C",
            end_type="num", end_value=15, end_color="FFC7CE"
        )
        ws.conditional_formatting.add(
            f"{col_letter}{header_row + 1}:{col_letter}{max_row}",
            color_scale_distance
        )

    # Пастельные цвета для Зоны
    zone_fills = {
        "Центр": PatternFill(start_color="E8F4F8", end_color="E8F4F8",
                             fill_type="solid"),
        "Срединная зона": PatternFill(start_color="F0F8E8",
                                      end_color="F0F8E8",
                                      fill_type="solid"),
        "Спальный район": PatternFill(start_color="FFF8E8",
                                      end_color="FFF8E8",
                                      fill_type="solid")
    }
    if col_zone:
        for row_num in range(header_row + 1, max_row + 1):
            z = ws.cell(row=row_num, column=col_zone).value
            if z in zone_fills:
                ws.cell(row=row_num, column=col_zone).fill = zone_fills[z]

    # Ширина столбцов
    widths = {
        "A": 20, "B": 22, "C": 25, "D": 35,
        "E": 18, "F": 16, "G": 18, "H": 20,
        "I": 22, "J": 15, "K": 18, "L": 8,
        "M": 12, "N": 15, "O": 10, "P": 10,
        "Q": 10, "R": 32, "S": 12, "T": 12,
        "U": 12, "V": 12
    }
    for col, w in widths.items():
        if col in ws.column_dimensions:
            ws.column_dimensions[col].width = w

    ws.row_dimensions[header_row].height = 25
    for r in range(header_row + 1, max_row + 1):
        ws.row_dimensions[r].height = 20


def main():
    print(f"Чтение файла: {INPUT_FILE}")
    wb_src = openpyxl.load_workbook(INPUT_FILE)
    ws_src = wb_src[SHEET_NAME]

    # Читаем все строки
    rows = list(ws_src.iter_rows(values_only=True))
    headers = rows[0]
    data_rows = rows[1:]
    total = len(data_rows)

    # Считаем по зонам
    col_zone_idx = headers.index("Зона") if "Зона" in headers else None
    zone_counts = Counter()
    if col_zone_idx is not None:
        for row in data_rows:
            if row[col_zone_idx]:
                zone_counts[row[col_zone_idx]] += 1

    parts = math.ceil(total / MAX_ROWS_PER_FILE)

    print(f"✓ Найдено компаний: {total}")
    print(f"✓ По зонам: {dict(zone_counts)}")
    print(f"✓ Будет файлов: {parts}\n")

    for part in range(parts):
        start = part * MAX_ROWS_PER_FILE
        end = min(start + MAX_ROWS_PER_FILE, total)
        chunk = data_rows[start:end]

        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        ws_new.title = SHEET_NAME

        # Добавляем заголовки
        ws_new.append(headers)
        # Добавляем данные
        for row in chunk:
            ws_new.append(list(row))

        # Оформление + гиперссылки + статистика
        style_and_links(ws_new, headers, zone_counts)

        out_name = f"Tablitsa_formatted_part{part+1}.xlsx"
        wb_new.save(out_name)
        print(f"✓ Сохранён: {out_name} (компании {start+1}–{end})")

    print("\n✅ Готово!")

if __name__ == "__main__":
    main()
